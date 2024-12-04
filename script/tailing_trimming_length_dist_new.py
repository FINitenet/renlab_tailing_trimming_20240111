#!/usr/bin/python3
# -*- coding:utf-8 -*-
# @FileName  :tailing_trimming_length_dist_new.py
# @Time      :2023/12/30 00:47:49
# @Author    :Yuchen@rlab
# @Description: This script is used to calculate the distribution of length of reads mapped to miRNA and 5GCM
# @version: 1

import os
import glob
import pysam
import argparse
from multiprocessing import Process
from collections import Counter
from collections import OrderedDict
from openpyxl import Workbook

def build_end_matix(meta_file):
    end = {}
    with open(meta_file, "r") as f:
        for line in f:
            inf = line.strip().split("\t")
            inf2 = inf[0].split("-")
            hairpin = inf2[0] + "-" + str(inf2[1])
            end[hairpin + "-" + inf[2]] = int(inf[2]) + len(inf[1]) - 1
    return end

def length_dist(samfile, end):
    lengthRead = Counter()
    length5GCM = Counter()
    total = 0
    readname = {}

    with pysam.AlignmentFile(samfile, "r") as f:
        for line in f:
            mirnaid = str(line.reference_name) + "-" + str(int(line.reference_start) + 1)
            queryid = line.query_name.split("-")[0] + "-" + line.query_name.split("-")[1]
            querycut = int(line.query_name.split("--")[1])

            lenRead = len(line.query_name.split("-")[2])
            end5GCM = line.reference_start + lenRead - querycut
            len5GCM = lenRead - querycut

            if mirnaid in end and queryid not in readname:
                if querycut == 0 and end5GCM > end[mirnaid]:
                    len5GCM = len(line.query_name.split("-")[2])

                lengthRead[lenRead] += 1
                length5GCM[len5GCM] += 1
                total += 1

            readname[queryid] = readname.get(queryid, 0) + 1

    # Create a new workbook
    wb = Workbook()

    # Create three worksheets
    sheet1 = wb.active
    sheet1.title = "total reads mapped to miRNA"
    sheet2 = wb.create_sheet("distribution of whole reads")
    sheet3 = wb.create_sheet("distribution of 5GMC")

    # Write values
    sheet1["A1"] = "total reads mapped to miRNA:"
    sheet1["B1"] = total

    lengthRead = OrderedDict(sorted(lengthRead.items(), key=lambda x: x[0]))
    length5GCM = OrderedDict(sorted(length5GCM.items(), key=lambda x: x[0]))

    # Populate sheet2 (distribution of whole reads)
    rows_sheet2 = [("Length", "Count", "Percentage")]
    rows_sheet2 += [(key, value, "{:.4f}".format(value / total).zfill(6)) for key, value in lengthRead.items()]
    for row in rows_sheet2:
        sheet2.append(row)

    # Populate sheet3 (distribution of 5GMC)
    rows_sheet3 = [("Length", "Count", "Percentage")]
    rows_sheet3 += [(key, value, "{:.4f}".format(value / total).zfill(6)) for key, value in length5GCM.items()]
    for row in rows_sheet3:
        sheet3.append(row)

    # Save the workbook
    output_file = os.path.join(OUTDIR, SAMPLE_NAME + "_len_dist.xlsx")
    wb.save(output_file)

##### Argument parsing / help message /version

parser = argparse.ArgumentParser(prog=os.path.basename(__file__))
parser.add_argument('-i', '--input', help='inputdir containing merged-alignment-sorted.sam', required=True)
parser.add_argument('-o', '--output', help='output path', required=True)
parser.add_argument('-f', help='miRNA meta file, contains miRNA start end length', 
                    default = "/bios-store1/chenyc/scripts/renlab_tailing_trimming_20240111/source/miRNA_start_sequence_length_new.txt")
parser.add_argument('-v', '--version', action='version', version='%(prog)s 20240307')

args = parser.parse_args()


##### Script control .. go through each file and process
if __name__ == "__main__":
    
    INPUTDIR = args.input
    OUTDIR = args.output
    METAFILE = args.f

    SAMPLE_NAMES = os.listdir(INPUTDIR)
    

    end = build_end_matix(METAFILE)

    process = []
    for SAMPLE_NAME in SAMPLE_NAMES:
        SAMFILE = os.path.join(INPUTDIR, SAMPLE_NAME, "merged-alignment-sorted.sam")
        print("Processing: ", SAMPLE_NAME)

        p = Process(target=length_dist, args=(SAMFILE, end))
        p.start()
        process.append(p)
    for p in process:
        p.join()

    print("All done!")